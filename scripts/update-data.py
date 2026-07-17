#!/usr/bin/env python3
"""
RASA Dashboard — Data Update Script
====================================
Run this script whenever new period Excel files are added to Financial-Dashboard/Data/.
It reads all Excel files, extracts every metric, and writes lib/dashboard-data.json.

Usage:
    python3 scripts/update-data.py

Requirements:
    pip install openpyxl
"""

import json
import os
import sys
import openpyxl
from pathlib import Path

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
BASE_DIR     = SCRIPT_DIR.parent
DATA_DIR     = BASE_DIR / 'Data'
JSON_PATH    = BASE_DIR / 'lib' / 'dashboard-data.json'

# ── Period index → filename ───────────────────────────────────────────────────
# Loaded dynamically from lib/periods-config.json so the admin UI and this
# script always stay in sync. No manual edits needed when a new period is added.
_cfg = json.loads((BASE_DIR / 'lib' / 'periods-config.json').read_text())
PERIOD_FILES = [(p['idx'], p['filename']) for p in _cfg]

# ── Sheet name → entity key ───────────────────────────────────────────────────
SHEET_TO_ENTITY = {
    'Consolidated':               'Consolidated',
    '1 - RASA Worldwide':         'RASA Worldwide',
    '2 - RASA - Ballpark':        'Ballpark',
    '4 - RASA - MVT':             'MVT',
    '5 - RASA - National Landin': 'National Landing',
    '6 - RASA - Mosaic':          'Mosaic',
    '7 - RASA - Rockville':       'Rockville',
}

# ── Excel column indices (0-based) ────────────────────────────────────────────
COL_LABEL      = 0   # Column A
COL_ACTUAL     = 1   # Column B
COL_BUDGET     = 4   # Column E
COL_PRIOR_YEAR = 10  # Column K

# ── Label mapping: Excel label → JSON key (only where they differ) ────────────
EXCEL_TO_JSON = {
    'ShareBite Commissions': 'Sharebite Commissions',
}

# ── All 244 metric keys expected in the JSON ──────────────────────────────────
ALL_KEYS = [
    # Sales
    'Food', 'N/A Beverage', 'Beer', 'Liquor', 'Wine', 'Retail',
    'Total Single Plate In-House Sales',
    'Takeout Sales - RASA App', 'Takeout Sales - DoorDash', 'Takeout Sales - GrubHub',
    'Takeout Sales - Uber Eats', 'Takeout Sales - Too Good To Go',
    'Takeout Sales - RASA Website', 'Takeout Sales - Ritual',
    'Total Takeout Sales',
    'Delivery Sales - DoorDash', 'Delivery Sales - GrubHub', 'Delivery Sales - Uber Eats',
    'Delivery Sales - RASA App', 'Delivery Sales - RASA Website',
    'Delivery Sales - Hungry Marketplace', 'Delivery Sales - Google',
    'Total Delivery Sales',
    'Catering Sales - RASA Website', 'Catering Sales - EZ Cater',
    'Catering Sales - Foodworks', 'Catering Sales - Cater Cow',
    'Catering Sales - Territory Foods', 'Catering Sales - Hungry Marketplace',
    'Catering Sales - Sharebite', 'Catering Sales - Cater 2 Me',
    'Catering Sales - ZeroCater', 'Catering Sales - Food Fleet',
    'Catering Sales - Foodee', 'Catering Sales - WCK',
    'Total Catering Sales - Other 3rd Party', 'Total Catering Sales',
    'Offsites - Fooda', 'Offsites - Aramark', 'Offsites - Eurest',
    'Offsites - Metz Corp', 'Offsites - Guest Services', 'Offsites - Cureate',
    'Offsites - Compass', 'Offsites - Taher',
    'Total Offsites',
    'Delivery Fee - Online Ordering', 'Delivery Fee - EZ Cater',
    'Delivery Fee - Fooda', 'Delivery Fee - Cater Cow',
    'Delivery Fee - Uber Eats', 'Delivery Fee - GrubHub',
    'Delivery Fee - Sharebite', 'Delivery Fee - ZeroCater',
    'Delivery Fee - Cater 2 Me', 'Delivery Fee - Compass',
    'Delivery Fee - Delivery by DoorDash (WL)',
    'Total Delivery Fee Income',
    'Delivery Service Charge',
    'Sales Adjustments', 'Open App Sales Adjustments', 'Open Tickets',
    'Squad Meals', 'Open App Loyalty', 'SLT Comps', 'LSM Comps',
    'Manager Comps', 'Catering Comps', 'Open App Store Credits',
    'Open App Promos', 'Squad Discount', 'Partnerships',
    'Discounts / Refunds Given Other',
    'Free Meal Cards (This Ones on Us)', 'Magic Fund',
    'Total Discounts / Refunds',
    'Total Sales',
    # COGS
    'COGS - Produce', 'COGS - Grocery Dry & Canned', 'COGS - Poultry',
    'COGS - Dairy', 'COGS - Frozen', 'COGS - Meat', 'COGS - Seafood',
    'COGS - Spices', 'COGS - Prepped Food',
    'Total Food Costs',
    'COGS - N/A Beverage', 'COGS - Beer', 'COGS - Liquor', 'COGS - Wine',
    'Total Beverage Costs',
    'COGS - Paper Supplies', 'COGS - Catering Supplies', 'COGS - Event Supplies',
    'Total Paper Costs',
    'Waste - Food', 'Waste - N/A Beverage', 'Waste - Beer',
    'Waste - Liquor', 'Waste - Wine',
    'Total Waste Costs',
    'COGS - Freight, Delivery, & Sales Tax',
    'COGS - Fuel Surcharge',
    'COGS - Credit for Discounts',
    'Total Cost of Goods Sold',
    'Gross Profit',
    # Labor
    'Management', 'Hourly', 'Overtime', 'Bonus',
    'FICA Tax Expense', 'FUTA Tax Expense', 'SUTA Tax Expense (VA)',
    'SUTA Tax Expense (DC)', 'FL Tax Expense (DC)', 'Other Payroll Taxes',
    'Total Payroll Taxes',
    'Payroll Processing Fee', 'Health Insurance',
    'Total Payroll Expenses',
    'Prime Profit',
    # OpEx — Supplies
    'Cleaning & Maintenance Supplies', 'Kitchen Supplies', 'FOH Supplies',
    'Towel & Doormat Service', 'Kitchen Smallwares', 'Uniforms', 'Office Supplies',
    'Total Supplies',
    # OpEx — Marketing
    'DoorDash Promos', 'EZ Cater Promos', 'GrubHub Promos', 'Uber Eats Promos',
    'Total Third-Party Promos',
    'Other Marketing Expenses',
    'Total Marketing',
    # OpEx — Delivery Fees
    'First Delivery Fees', 'Nash Fees', 'Open App Fees',
    'Total Delivery Fees',
    # OpEx — 3rd Party Fees / Commissions
    'Cater Cow Commissions', 'DoorDash Commissions', 'EZ Cater Commissions',
    'Fooda Commissions', 'GrubHub Commissions', 'Too Good To Go Commissions',
    'Uber Eats Commissions', 'Sharebite Commissions',
    'Aramark Commissions', 'Eurest Commissions', 'Metz Corp Commissions',
    'Guest Services Commissions', 'Foodworks Commissions',
    'Total Delivery Partner Commissions',
    'Total Third-Party Fees',
    # OpEx — Credit Card
    'Credit Card Fees', 'Chargebacks',
    'Total Credit Card Fees',
    # OpEx — Insurance
    'Workers Compensation Insurance', 'Business Owners Policy', 'Auto Insurance',
    'Total Insurance',
    # OpEx — R&M
    'Scrubber Expense', 'R&M - Hood Cleaning', 'R&M - Pest Control',
    'R&M - Building, General', 'R&M - Equipment, General', 'R&M - Refrigeration',
    'R&M - HVAC', 'R&M - Preventative Maintenance',
    'Total Repairs & Maintenance',
    # OpEx — Other
    'Equipment Rental', 'Bank Fees', 'Employee Incentives (M&E)',
    'Ground Transportation (Auto Expense)', 'Parking Fees', 'Tangible Property Tax',
    'Equipment Lease', 'Phone/Internet', 'POS & Technology',
    'Cost for Discounts', 'Licenses & Permits', 'Cash Over/Short',
    'Total Other Expenses',
    'Total Operating Expense',
    # Occupancy
    'Rent Expense', 'Common Area Maintenance (CAM)', 'Public Space Rental',
    'Insurance - Building', 'Real Estate Tax', 'Rent - 4/4/5',
    'Percentage Rent Expense',
    'Electric', 'Gas', 'Water/Sewer', 'Trash',
    'Total Utilities',
    'Total Occupancy Cost',
    'Store Level Profit',
    # Non-Controllable
    'Total Non Controllable Expense',
    # Corporate
    'Corp - Salaries & Wages', 'Corp - Payroll Taxes', 'Corp - Payroll Processing Fee',
    'Corp - Bonus', 'Corp - Health Insurance',
    'Total Corp - Payroll Expenses',
    'Corp - Parking Fees', 'Corp - Ground Transportation', 'Corp - M&E',
    'Corp - Lodging', 'Corp - Airfare', 'Corp - Travel (New Market)',
    'Total Corp - Meals, Entertainment, & Travel',
    'Corp - Marketing', 'Corp - Marketing Food Drops',
    'Corp - 3rd Party Marketing Fees',
    'Corp - DoorDash Marketing Fees', 'Corp - EZ Cater Marketing Fees',
    'Corp - GrubHub Marketing Fees', 'Corp - Uber Eats Marketing Fees',
    'Total Corp - 3rd Party Marketing Fees',
    'Corp - Paid Media', 'Corp - SMS Marketing',
    'Total Corp - Marketing',
    'Corp - Insurance Expense', 'Corp - IT & Technology',
    'Corp - Accounting Fees', 'Corp - Legal Fees', 'Corp - Other Professional Fees',
    'Corp - Recruiting Fees', 'Corp - Dues & Subscriptions', 'Corp - Office Supplies',
    'Corp - Other Taxes', 'Corp - Rent Expense', 'Corp - Charitable Contributions',
    'Corp - Other Employee Incentives', 'Corp - R&D', 'Corp - Bank Fees',
    'Corp - Licenses & Permits',
    'Total Corporate Overhead & Other',
    # P&L summary
    'EBITDA', 'Net Income',
]

# Build reverse mapping: JSON key → Excel label (most are identical)
JSON_TO_EXCEL = {v: k for k, v in EXCEL_TO_JSON.items()}

N_PERIODS = len(PERIOD_FILES)
ENTITIES  = list(SHEET_TO_ENTITY.values())


def extract_sheet(ws, json_keys_set):
    """
    Scan worksheet rows. Return dict: json_key → (actual, budget, prior_year).
    Takes the first data row (actual not None) for each key.
    Section headers (actual=None) are skipped.
    """
    result = {}
    for row in ws.iter_rows(min_row=1, max_row=3000, min_col=1, max_col=11, values_only=True):
        excel_label = row[COL_LABEL]
        if not excel_label:
            continue
        actual = row[COL_ACTUAL] if len(row) > COL_ACTUAL else None
        if actual is None:
            continue  # section header — skip

        # Resolve JSON key
        excel_label_str = str(excel_label)
        if excel_label_str in EXCEL_TO_JSON:
            json_key = EXCEL_TO_JSON[excel_label_str]
        elif excel_label_str in json_keys_set:
            json_key = excel_label_str
        else:
            continue  # not a key we care about

        if json_key in result:
            continue  # already captured first occurrence

        budget    = row[COL_BUDGET]    if len(row) > COL_BUDGET    else None
        prior_yr  = row[COL_PRIOR_YEAR] if len(row) > COL_PRIOR_YEAR else None
        result[json_key] = (
            float(actual),
            float(budget)   if budget   is not None else 0.0,
            float(prior_yr) if prior_yr is not None else 0.0,
        )
    return result


def main():
    print("RASA Dashboard — Data Update")
    print("=" * 50)

    # Validate paths
    if not DATA_DIR.exists():
        print(f"ERROR: Data directory not found: {DATA_DIR}")
        sys.exit(1)
    if not JSON_PATH.exists():
        print(f"ERROR: JSON not found: {JSON_PATH}")
        sys.exit(1)

    # Load existing JSON to preserve metadata (periods, period_keys)
    with open(JSON_PATH) as f:
        data = json.load(f)

    periods     = data.get('periods', [])
    period_keys = data.get('period_keys', [])
    json_keys_set = set(ALL_KEYS)

    # Initialize fresh t12 structure — zero arrays for every key/entity
    t12 = {entity: {key: {'v': [0.0]*N_PERIODS, 'b': [0.0]*N_PERIODS, 'py': [0.0]*N_PERIODS}
                    for key in ALL_KEYS}
           for entity in ENTITIES}

    missing_files   = []
    missing_sheets  = []
    zero_key_counts = {}  # key → count of periods where it was found

    for period_idx, fname in PERIOD_FILES:
        fpath = DATA_DIR / fname
        if not fpath.exists():
            print(f"  WARNING: file not found — {fname}")
            missing_files.append(fname)
            continue

        print(f"  Reading {fname} (P index {period_idx}) ...", end=' ', flush=True)
        wb = openpyxl.load_workbook(str(fpath), data_only=True)

        for sheet_name, entity in SHEET_TO_ENTITY.items():
            if sheet_name not in wb.sheetnames:
                missing_sheets.append((fname, sheet_name))
                continue

            ws = wb[sheet_name]
            vals = extract_sheet(ws, json_keys_set)

            for json_key, (v, b, py) in vals.items():
                t12[entity][json_key]['v'][period_idx] = v
                t12[entity][json_key]['b'][period_idx] = b
                t12[entity][json_key]['py'][period_idx] = py
                zero_key_counts[json_key] = zero_key_counts.get(json_key, 0) + 1

        print("done")

    # Write JSON
    data['periods']     = periods
    data['period_keys'] = period_keys
    data['t12']         = t12

    with open(JSON_PATH, 'w') as f:
        json.dump(data, f, separators=(',', ':'))

    print()
    print(f"✓  Written to {JSON_PATH}")
    print(f"   Entities : {len(ENTITIES)}")
    print(f"   Periods  : {N_PERIODS}")
    print(f"   Keys     : {len(ALL_KEYS)}")

    # Report any keys that were never found in any file
    never_found = [k for k in ALL_KEYS if k not in zero_key_counts]
    if never_found:
        print(f"\n  ⚠  {len(never_found)} keys had no data in any Excel file:")
        for k in never_found:
            print(f"      {k}")

    if missing_files:
        print(f"\n  ⚠  {len(missing_files)} period files were missing:")
        for f in missing_files:
            print(f"      {f}")

    if missing_sheets:
        print(f"\n  ⚠  {len(missing_sheets)} sheets were missing:")
        for fname, sheet in missing_sheets:
            print(f"      {fname} → {sheet}")

    print()
    print("Done. Restart the Next.js dev server (or redeploy) to pick up new data.")


if __name__ == '__main__':
    main()
